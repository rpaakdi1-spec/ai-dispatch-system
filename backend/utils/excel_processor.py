import pandas as pd
from typing import List, Dict, Any
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io


class ExcelProcessor:
    """Excel file processing utilities"""
    
    @staticmethod
    def read_excel(file_path: str, sheet_name: str = None) -> pd.DataFrame:
        """Read Excel file into pandas DataFrame"""
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name or 0)
            return df
        except Exception as e:
            raise ValueError(f"Failed to read Excel file: {str(e)}")
    
    @staticmethod
    def read_excel_from_bytes(file_bytes: bytes, sheet_name: str = None) -> pd.DataFrame:
        """Read Excel file from bytes into pandas DataFrame"""
        try:
            df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name or 0)
            return df
        except Exception as e:
            raise ValueError(f"Failed to read Excel file: {str(e)}")
    
    @staticmethod
    def validate_columns(df: pd.DataFrame, required_columns: List[str]) -> tuple[bool, List[str]]:
        """Validate that DataFrame has all required columns"""
        missing_columns = [col for col in required_columns if col not in df.columns]
        return len(missing_columns) == 0, missing_columns
    
    @staticmethod
    def parse_clients_excel(df: pd.DataFrame) -> List[Dict[str, Any]]:
        """
        Parse clients data from Excel DataFrame
        
        Expected columns:
        - 거래처코드, 거래처명, 구분 (상차/하차/양쪽), 주소
        - 상차가능시작, 상차가능종료, 하차가능시작, 하차가능종료
        - 지게차유무 (Y/N), 대형차진입 (Y/N)
        """
        required_columns = ['거래처코드', '거래처명', '구분', '주소']
        is_valid, missing = ExcelProcessor.validate_columns(df, required_columns)
        
        if not is_valid:
            raise ValueError(f"Missing required columns: {', '.join(missing)}")
        
        clients = []
        
        for idx, row in df.iterrows():
            # Parse service type
            service_type = row['구분'].strip()
            if '상차' in service_type and '하차' in service_type:
                service_type_en = 'both'
            elif '상차' in service_type:
                service_type_en = 'pickup'
            elif '하차' in service_type:
                service_type_en = 'delivery'
            else:
                service_type_en = 'both'
            
            client = {
                'client_id': str(row['거래처코드']).strip(),
                'client_name': str(row['거래처명']).strip(),
                'service_type': service_type_en,
                'address': str(row['주소']).strip(),
                'pickup_time_start': str(row.get('상차가능시작', '09:00')).strip() if pd.notna(row.get('상차가능시작')) else '09:00',
                'pickup_time_end': str(row.get('상차가능종료', '17:00')).strip() if pd.notna(row.get('상차가능종료')) else '17:00',
                'delivery_time_start': str(row.get('하차가능시작', '09:00')).strip() if pd.notna(row.get('하차가능시작')) else '09:00',
                'delivery_time_end': str(row.get('하차가능종료', '17:00')).strip() if pd.notna(row.get('하차가능종료')) else '17:00',
                'has_forklift': str(row.get('지게차유무', 'N')).upper().startswith('Y') if pd.notna(row.get('지게차유무')) else False,
                'allows_large_truck': str(row.get('대형차진입', 'Y')).upper().startswith('Y') if pd.notna(row.get('대형차진입')) else True,
                'contact_phone': str(row.get('연락처', '')).strip() if pd.notna(row.get('연락처')) else None,
            }
            
            clients.append(client)
        
        return clients
    
    @staticmethod
    def parse_orders_excel(df: pd.DataFrame) -> List[Dict[str, Any]]:
        """
        Parse orders data from Excel DataFrame
        
        Expected columns:
        - 주문번호, 상차거래처코드, 하차거래처코드
        - 온도대 (냉동/냉장/상온), 팔레트수, 중량(kg)
        - 상차시작시간, 상차종료시간, 하차시작시간, 하차종료시간
        """
        required_columns = ['주문번호', '상차거래처코드', '하차거래처코드', '온도대', '팔레트수', '중량(kg)']
        is_valid, missing = ExcelProcessor.validate_columns(df, required_columns)
        
        if not is_valid:
            raise ValueError(f"Missing required columns: {', '.join(missing)}")
        
        orders = []
        
        for idx, row in df.iterrows():
            # Parse temperature type
            temp_type = str(row['온도대']).strip()
            if '냉동' in temp_type:
                temp_type_en = 'frozen'
            elif '냉장' in temp_type:
                temp_type_en = 'chilled'
            else:
                temp_type_en = 'ambient'
            
            order = {
                'order_id': str(row['주문번호']).strip(),
                'pickup_client_id': str(row['상차거래처코드']).strip(),
                'delivery_client_id': str(row['하차거래처코드']).strip(),
                'temperature_type': temp_type_en,
                'required_pallets': int(row['팔레트수']),
                'weight_kg': float(row['중량(kg)']),
                'pickup_time_start': str(row.get('상차시작시간', '09:00')).strip() if pd.notna(row.get('상차시작시간')) else '09:00',
                'pickup_time_end': str(row.get('상차종료시간', '11:00')).strip() if pd.notna(row.get('상차종료시간')) else '11:00',
                'delivery_time_start': str(row.get('하차시작시간', '13:00')).strip() if pd.notna(row.get('하차시작시간')) else '13:00',
                'delivery_time_end': str(row.get('하차종료시간', '17:00')).strip() if pd.notna(row.get('하차종료시간')) else '17:00',
                'priority': 'normal',
                'status': 'pending',
            }
            
            orders.append(order)
        
        return orders
    
    @staticmethod
    def create_template_excel(template_type: str, output_path: str):
        """Create Excel template file"""
        wb = Workbook()
        ws = wb.active
        
        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        if template_type == "clients":
            ws.title = "거래처마스터"
            headers = ['거래처코드', '거래처명', '구분', '주소', '상세주소', '연락처', 
                      '상차가능시작', '상차가능종료', '하차가능시작', '하차가능종료',
                      '지게차유무', '대형차진입', '비고']
            example_data = ['CUST-0001', '(주)서울냉동', '양쪽', '서울 송파구 문정동 123', '', '02-1234-5678',
                          '09:00', '17:00', '09:00', '17:00', 'Y', 'Y', '']
            
        elif template_type == "orders":
            ws.title = "주문목록"
            headers = ['주문번호', '상차거래처코드', '하차거래처코드', '온도대', 
                      '팔레트수', '중량(kg)', '상차시작시간', '상차종료시간',
                      '하차시작시간', '하차종료시간', '우선순위', '비고']
            example_data = ['ORD-001', 'CUST-0001', 'CUST-0100', '냉동', 
                          6, 3000, '09:00', '11:00', '13:00', '17:00', '보통', '']
            
        elif template_type == "vehicles":
            ws.title = "차량마스터"
            headers = ['차량코드', 'UVIS단말기ID', '차량번호', '차량타입', '톤수',
                      '최대팔레트', '최대중량(kg)', '최저온도', '최고온도', '겸용차량',
                      '기사명', '기사연락처', '비고']
            example_data = ['TRUCK-001', 'UVIS-DVC-12345', '12가3456', '냉동', 5.0,
                          16, 10000, -25.0, -15.0, 'N', '홍길동', '010-1234-5678', '']
        else:
            raise ValueError(f"Unknown template type: {template_type}")
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write example data
        for col, value in enumerate(example_data, 1):
            ws.cell(row=2, column=col, value=value)
        
        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        wb.save(output_path)
